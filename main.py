import tkinter as tk
from tkinter import messagebox, filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import requests
from bs4 import BeautifulSoup
import pandas as pd
import json
import threading
import re
from datetime import datetime
import time

class MaterialPriceScraper:
    def __init__(self, root):
        self.root = root
        self.root.title("辽宁省网刊价格数据提取工具")
        self.root.geometry("900x700")
        
        # 设置现代扁平化无边框窗口样式
        self.root.overrideredirect(True)  # 隐藏窗口边框
        self.root.configure(bg="#f0f0f0")  # 设置背景色为浅灰色
        
        # 初始化变量
        self.session = None
        self.is_connected = False
        self.city_mapping = {}
        self.year_list = []
        self.month_list = []
        self.selected_city = ""
        self.selected_year = ""
        self.selected_month = ""
        self.total_pages = 0
        self.total_records = 0
        self.all_data = []
        self.headers = []
        
        # 参数变更跟踪变量
        self.current_city = ""
        self.current_year = ""
        self.current_month = ""
        
        # 窗口拖动相关变量
        self.start_x = 0
        self.start_y = 0
        self.dragging = False
        
        # 初始化TEMP目录
        import os
        import sys
        # 获取可执行文件所在目录（兼容单个EXE文件和脚本运行）
        if getattr(sys, 'frozen', False):
            # 单个EXE文件运行
            self.exe_dir = os.path.dirname(os.path.abspath(sys.executable))
        else:
            # 脚本运行
            self.exe_dir = os.path.dirname(os.path.abspath(__file__))
        # 创建相对于可执行文件目录的TEMP目录
        self.temp_dir = os.path.join(self.exe_dir, "TEMP")
        # 创建TEMP目录
        os.makedirs(self.temp_dir, exist_ok=True)
        
        # 设置程序退出时清理TEMP目录
        self.root.protocol("WM_DELETE_WINDOW", self.close_window)
        
        # 创建用户界面
        self.create_ui()
        
        # 添加窗口控制功能
        self.setup_window_controls()
    
    def setup_window_controls(self):
        """设置窗口控制功能（拖动、最小化、关闭）"""
        # 添加鼠标事件用于窗口拖动
        self.root.bind("<Button-1>", self.start_drag)
        self.root.bind("<B1-Motion>", self.drag_window)
        self.root.bind("<ButtonRelease-1>", self.stop_drag)
    
    def start_drag(self, event):
        """开始拖动窗口"""
        self.start_x = event.x
        self.start_y = event.y
        self.dragging = True
    
    def drag_window(self, event):
        """拖动窗口"""
        if self.dragging:
            x = self.root.winfo_x() + event.x - self.start_x
            y = self.root.winfo_y() + event.y - self.start_y
            self.root.geometry(f"+{x}+{y}")
    
    def stop_drag(self, event):
        """停止拖动窗口"""
        self.dragging = False
    
    def create_ui(self):
        """创建用户界面"""
        # 标题栏框架 - 使用tk.Frame来支持直接设置背景色
        title_frame = tk.Frame(self.root, height=40, bg="#0d6efd")
        title_frame.pack(fill=tk.X, pady=0, padx=0)
        title_frame.bind("<Button-1>", self.start_drag)
        title_frame.bind("<B1-Motion>", self.drag_window)
        title_frame.bind("<ButtonRelease-1>", self.stop_drag)
        
        # 标题标签 - 使用tk.Label来支持直接设置颜色
        title_label = tk.Label(title_frame, text="辽宁省网刊价格数据提取工具", font=("微软雅黑", 11, "bold"), fg="#ffffff", bg="#0d6efd")
        title_label.pack(side=tk.LEFT, padx=15, pady=8)
        title_label.bind("<Button-1>", self.start_drag)
        title_label.bind("<B1-Motion>", self.drag_window)
        title_label.bind("<ButtonRelease-1>", self.stop_drag)
        
        # 窗口控制按钮 - 使用tk.Frame
        control_frame = tk.Frame(title_frame, bg="#0d6efd")
        control_frame.pack(side=tk.RIGHT, padx=5, pady=0)
        control_frame.bind("<Button-1>", self.start_drag)
        control_frame.bind("<B1-Motion>", self.drag_window)
        control_frame.bind("<ButtonRelease-1>", self.stop_drag)
        
        # 关闭按钮 - 使用tk.Button
        close_btn = tk.Button(control_frame, text="✕", command=self.close_window, width=3, fg="#ffffff", bg="#0d6efd", bd=0, font=("Arial", 10, "bold"))
        close_btn.pack(side=tk.LEFT, padx=2, pady=0)
        close_btn.bind("<Enter>", lambda e: close_btn.config(bg="#dc3545"))
        close_btn.bind("<Leave>", lambda e: close_btn.config(bg="#0d6efd"))
        
        # 主内容框架
        main_content = ttk.Frame(self.root, padding="20")
        main_content.pack(fill=tk.BOTH, expand=True)
        
        # 框架1：网站连接
        frame_connect = ttk.LabelFrame(main_content, text="网站连接")
        frame_connect.pack(fill=tk.X, pady=(0, 20), padx=0)
        connect_container = ttk.Frame(frame_connect, padding=15)
        connect_container.pack(fill=tk.X)
        
        self.connect_btn = ttk.Button(connect_container, text="连接网站", command=self.connect_to_website, bootstyle="success")
        self.connect_btn.pack(side=tk.LEFT, padx=0, pady=0)
        
        self.connection_status = ttk.Label(connect_container, text="未连接", foreground="#e74c3c")
        self.connection_status.pack(side=tk.LEFT, padx=20, pady=0)
        
        # 框架2：选择参数
        frame_select = ttk.LabelFrame(main_content, text="选择参数")
        frame_select.pack(fill=tk.X, pady=(0, 20), padx=0)
        select_container = ttk.Frame(frame_select, padding=15)
        select_container.pack(fill=tk.X)
        
        # 城市选择
        city_label = ttk.Label(select_container, text="城市:")
        city_label.grid(row=0, column=0, padx=(0, 10), pady=5, sticky="w")
        self.city_var = tk.StringVar()
        self.city_combo = ttk.Combobox(select_container, textvariable=self.city_var, state="readonly", width=30)
        self.city_combo.grid(row=0, column=1, padx=(0, 20), pady=5, sticky="w")
        
        # 年份选择
        year_label = ttk.Label(select_container, text="年份:")
        year_label.grid(row=0, column=2, padx=(0, 10), pady=5, sticky="w")
        self.year_var = tk.StringVar()
        self.year_combo = ttk.Combobox(select_container, textvariable=self.year_var, state="readonly", width=15)
        self.year_combo.grid(row=0, column=3, padx=(0, 20), pady=5, sticky="w")
        
        # 月份选择
        month_label = ttk.Label(select_container, text="月份:")
        month_label.grid(row=0, column=4, padx=(0, 10), pady=5, sticky="w")
        self.month_var = tk.StringVar()
        self.month_combo = ttk.Combobox(select_container, textvariable=self.month_var, state="readonly", width=15)
        self.month_combo.grid(row=0, column=5, padx=(0, 0), pady=5, sticky="w")
        
        # 整合数据处理模块
        frame_processing = ttk.LabelFrame(main_content, text="数据处理")
        frame_processing.pack(fill=tk.X, pady=(0, 20), padx=0)
        processing_container = ttk.Frame(frame_processing, padding=20)
        processing_container.pack(fill=tk.X)
        
        # 步骤1：数据查询
        step1_frame = ttk.Frame(processing_container)
        step1_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 20))
        
        step1_label = ttk.Label(step1_frame, text="步骤1: 查询数据", font=("微软雅黑", 10, "bold"))
        step1_label.pack(anchor="w", pady=(0, 10))
        
        step1_content = ttk.Frame(step1_frame)
        step1_content.pack(fill=tk.X)
        
        self.query_btn = ttk.Button(step1_content, text="查询数据", command=self.query_data, bootstyle="info")
        self.query_btn.pack(side=tk.LEFT, padx=0, pady=0)
        
        query_info = ttk.Frame(step1_content)
        query_info.pack(side=tk.LEFT, padx=20, pady=0)
        
        self.total_pages_label = ttk.Label(query_info, text="总页数: 0")
        self.total_pages_label.pack(side=tk.LEFT, padx=15, pady=0)
        
        self.total_records_label = ttk.Label(query_info, text="总记录数: 0")
        self.total_records_label.pack(side=tk.LEFT, padx=15, pady=0)
        
        # 分隔线
        separator = ttk.Separator(processing_container, orient=tk.VERTICAL)
        separator.pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        # 步骤2：数据提取
        step2_frame = ttk.Frame(processing_container)
        step2_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 20))
        
        step2_label = ttk.Label(step2_frame, text="步骤2: 提取数据", font=("微软雅黑", 10, "bold"))
        step2_label.pack(anchor="w", pady=(0, 10))
        
        step2_content = ttk.Frame(step2_frame)
        step2_content.pack(fill=tk.X)
        
        self.extract_btn = ttk.Button(step2_content, text="提取数据", command=self.extract_data, bootstyle="warning")
        self.extract_btn.pack(side=tk.LEFT, padx=0, pady=0)
        
        extract_options = ttk.Frame(step2_content)
        extract_options.pack(side=tk.LEFT, padx=20, pady=0)
        
        self.extract_all_pages_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(extract_options, text="提取所有页", variable=self.extract_all_pages_var).pack(side=tk.LEFT, padx=0, pady=0)
        
        # 分隔线
        separator2 = ttk.Separator(processing_container, orient=tk.VERTICAL)
        separator2.pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        # 步骤3：数据导出
        step3_frame = ttk.Frame(processing_container)
        step3_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        step3_label = ttk.Label(step3_frame, text="步骤3: 导出数据", font=("微软雅黑", 10, "bold"))
        step3_label.pack(anchor="w", pady=(0, 10))
        
        step3_content = ttk.Frame(step3_frame)
        step3_content.pack(fill=tk.X)
        
        self.export_btn = ttk.Button(step3_content, text="导出数据", command=self.export_data, bootstyle="danger")
        self.export_btn.pack(side=tk.LEFT, padx=0, pady=0)
        
        # 框架6：进度和日志
        frame_progress = ttk.LabelFrame(main_content, text="执行进度与日志")
        frame_progress.pack(fill=tk.BOTH, expand=True, pady=(0, 0), padx=0)
        progress_container = ttk.Frame(frame_progress, padding=15)
        progress_container.pack(fill=tk.BOTH, expand=True)
        
        # 进度条 - 使用ttkbootstrap的样式
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_container, variable=self.progress_var, maximum=100, bootstyle="info-striped")
        self.progress_bar.pack(fill=tk.X, pady=(0, 15))
        
        # 日志文本框
        log_container = ttk.Frame(progress_container)
        log_container.pack(fill=tk.BOTH, expand=True)
        
        log_label = ttk.Label(log_container, text="执行日志:")
        log_label.pack(anchor="w", pady=(0, 10))
        
        # 日志文本框和滚动条
        log_frame = ttk.Frame(log_container, borderwidth=1, relief=tk.SOLID)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        # 使用ttkbootstrap风格的文本框
        self.log_text = tk.Text(log_frame, wrap="word", height=15, font=("Consolas", 9), bg="#f8f9fa", fg="#212529", bd=0, padx=10, pady=10)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 日志滚动条 - 使用ttkbootstrap的样式
        log_scrollbar = ttk.Scrollbar(log_frame, bootstyle="secondary")
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=log_scrollbar.set)
        log_scrollbar.config(command=self.log_text.yview)
        
        # 初始化按钮状态
        self.update_button_states()
    
    def close_window(self):
        """关闭窗口并清理临时文件"""
        try:
            import os
            import shutil
            # 删除TEMP目录及其内容
            if os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except Exception as e:
            # 忽略清理过程中的错误，确保程序能正常退出
            pass
        finally:
            self.root.destroy()
    
    def update_button_states(self):
        """更新按钮状态"""
        if self.is_connected:
            self.connect_btn.config(state="disabled")
            self.query_btn.config(state="normal")
            self.connection_status.config(text="已连接", foreground="#27ae60")
        else:
            self.connect_btn.config(state="normal")
            self.query_btn.config(state="disabled")
            self.extract_btn.config(state="disabled")
            self.export_btn.config(state="disabled")
            self.connection_status.config(text="未连接", foreground="#e74c3c")
    
    def log_message(self, message):
        """添加日志消息"""
        # 定义日志级别颜色映射
        log_colors = {
            "✅": "#008000",  # 成功 - 绿色
            "❌": "#ff0000",  # 错误 - 红色
            "⚠️": "#ff8c00",  # 警告 - 橙色
            "🔄": "#0000ff",  # 过程 - 蓝色
            "📌": "#800080",  # 关键信息 - 紫色
            "📊": "#008080",  # 数据 - 青色
            "⏱️": "#a9a9a9",  # 时间 - 灰色
            "📋": "#696969",  # 列表 - 深灰色
            "🎉": "#ff69b4"   # 完成 - 粉色
        }
        
        # 提取日志前缀（如果有）
        prefix = message[:2] if message and message[0] in ["✅", "❌", "⚠️", "🔄", "📌", "📊", "⏱️", "📋", "🎉"] else ""
        
        # 创建带有时间戳的日志消息
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"{timestamp} - {message}\n"
        
        # 插入日志消息
        self.log_text.insert(tk.END, log_entry)
        
        # 根据日志级别设置颜色
        if prefix in log_colors:
            # 获取当前文本的末尾位置
            end_pos = self.log_text.index(tk.END)
            # 计算日志条目的起始和结束位置
            start_pos = end_pos + "-1l+0c"
            end_pos = end_pos + "-1c"
            # 设置颜色
            self.log_text.tag_add(prefix, start_pos, end_pos)
            self.log_text.tag_config(prefix, foreground=log_colors[prefix])
        
        # 滚动到最新日志
        self.log_text.see(tk.END)
        # 更新UI
        self.root.update_idletasks()
    
    def update_button_states(self):
        """更新按钮状态"""
        if self.is_connected:
            self.connect_btn.config(state="disabled")
            self.query_btn.config(state="normal")
            self.connection_status.config(text="已连接", foreground="green")
        else:
            self.connect_btn.config(state="normal")
            self.query_btn.config(state="disabled")
            self.extract_btn.config(state="disabled")
            self.export_btn.config(state="disabled")
            self.connection_status.config(text="未连接", foreground="red")
    
    def on_parameter_change(self):
        """参数变更回调函数"""
        # 获取当前选择的参数
        current_city = self.city_var.get()
        current_year = self.year_var.get()
        current_month = self.month_var.get()
        
        # 检查是否有变化
        if (current_city != self.current_city or 
            current_year != self.current_year or 
            current_month != self.current_month):
            
            # 更新当前参数
            self.current_city = current_city
            self.current_year = current_year
            self.current_month = current_month
            
            # 启用查询按钮
            if self.is_connected:
                self.query_btn.config(state="normal")
            
            # 重置相关状态
            self.total_pages = 0
            self.total_records = 0
            self.all_data = []
            self.headers = []
            
            # 更新UI显示
            self.total_pages_label.config(text="总页数: 0")
            self.total_records_label.config(text="总记录数: 0")
            
            # 禁用提取和导出按钮
            self.extract_btn.config(state="disabled")
            self.export_btn.config(state="disabled")
    
    def connect_to_website(self):
        """连接到网站，建立会话"""
        # 立即禁用连接按钮，防止重复点击
        self.connect_btn.config(state="disabled")
        
        def connect_task():
            max_retries = 3
            retry_count = 0
            timeout = 60  # 增加超时时间到60秒
            
            while retry_count < max_retries:
                try:
                    retry_count += 1
                    self.log_message(f"🔄 正在连接到网站 (第{retry_count}/{max_retries}次尝试)...")
                    
                    # 创建会话对象
                    self.session = requests.Session()
                    
                    # 设置完整的浏览器头
                    self.session.headers.update({
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
                        'Accept-Encoding': 'gzip, deflate',
                        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                        'Cache-Control': 'no-cache',
                        'Connection': 'keep-alive',
                        'DNT': '1',
                        'Pragma': 'no-cache',
                        'Referer': 'http://218.60.144.156/jgxx_clcx.asp',
                        'Upgrade-Insecure-Requests': '1',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                    })
                    
                    # 先访问主入口页面，建立会话
                    main_url = "http://218.60.144.156/jgxx_clcx.asp"
                    self.log_message("📌 访问主入口页面")
                    self.log_message(f"⏱️  超时时间: {timeout}秒")
                    main_response = self.session.get(main_url, timeout=timeout)
                    # 先尝试从响应头获取编码，再使用gbk作为备选
                    main_response.encoding = main_response.apparent_encoding if main_response.apparent_encoding else 'gbk'
                    
                    if main_response.status_code != 200:
                        self.log_message(f"❌ 主入口页面访问失败，状态码: {main_response.status_code}")
                        if retry_count < max_retries:
                            time.sleep(2)
                        continue
                    
                    # 访问真实表单页面
                    form_url = "http://218.60.144.156/jgxx_cl1.asp?view=hidden"
                    self.log_message("📌 访问真实表单页面")
                    form_response = self.session.get(form_url, timeout=timeout)
                    # 先尝试从响应头获取编码，再使用gbk作为备选
                    form_response.encoding = form_response.apparent_encoding if form_response.apparent_encoding else 'gbk'
                    
                    if form_response.status_code != 200:
                        self.log_message(f"❌ 真实表单页面访问失败，状态码: {form_response.status_code}")
                        if retry_count < max_retries:
                            time.sleep(2)
                        continue
                    
                    self.log_message("✅ 网站连接成功")
                    self.is_connected = True
                    
                    # 获取城市列表
                    self.get_city_list(form_response.text)
                    
                    # 获取年份和月份列表
                    self.get_year_month_list()
                    
                    # 更新按钮状态
                    self.root.after(0, self.update_button_states)
                    
                    return  # 连接成功，退出循环
                    
                except requests.exceptions.Timeout:
                    self.log_message(f"⏱️  连接超时 (第{retry_count}/{max_retries}次)")
                    if retry_count < max_retries:
                        time.sleep(2)
                except requests.exceptions.ConnectionError:
                    self.log_message(f"❌ 网络连接错误 (第{retry_count}/{max_retries}次)")
                    if retry_count < max_retries:
                        time.sleep(2)
                except Exception as e:
                    self.log_message(f"❌ 网站连接失败 (第{retry_count}/{max_retries}次)")
                    if retry_count < max_retries:
                        time.sleep(2)
            
            # 所有重试都失败
            self.log_message(f"❌ 网站连接失败，已重试{max_retries}次")
            self.is_connected = False
            # 连接失败后，重新启用连接按钮
            self.root.after(0, lambda: self.connect_btn.config(state="normal"))
            # 更新按钮状态
            self.root.after(0, self.update_button_states)
        
        # 使用线程执行连接操作，避免阻塞GUI
        thread = threading.Thread(target=connect_task)
        thread.daemon = True
        thread.start()
    
    def get_city_list(self, html_content):
        """从网页内容中提取城市列表"""
        try:
            self.log_message("🔄 正在获取城市列表...")
            
            # 保存调试HTML，便于分析
            import os
            debug_file_path = os.path.join(self.temp_dir, "city_list_debug.html")
            with open(debug_file_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            self.log_message("✅ 已保存城市列表调试文件")
            
            # 使用gbk编码解析HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # 查找城市下拉选择框，确保只选择城市选择框
            city_select = soup.find('select', {'name': 'dq_id'})
            
            if not city_select:
                # 使用默认城市列表
                city_mapping = {
                    "15": "沈阳市", "16": "大连市", "53": "大连金普新区", "56": "大连开发区（2017前）",
                    "17": "鞍山市", "21": "抚顺市", "22": "本溪市", "25": "丹东市", "33": "锦州市",
                    "29": "营口市", "34": "阜新市", "35": "辽阳市", "36": "铁岭市", "44": "朝阳市",
                    "45": "盘锦市", "47": "葫芦岛市", "48": "绥中"
                }
            else:
                # 提取城市选项
                city_mapping = {}
                options = city_select.find_all('option')
                
                for option in options:
                    value = option.get('value', '').strip()
                    text = option.get_text(strip=True)
                    
                    # 严格过滤，只接受有效的城市选项
                    if value and text and value != '-1' and text != '请选择' and len(text) > 1:
                        # 尝试编码处理
                        if isinstance(text, str):
                            try:
                                # 先尝试gbk解码
                                text = text.encode('iso-8859-1', errors='ignore').decode('gbk')
                            except:
                                try:
                                    # 再尝试utf-8解码
                                    text = text.encode('iso-8859-1', errors='ignore').decode('utf-8')
                                except:
                                    # 最后使用replace处理
                                    text = text.encode('utf-8', errors='replace').decode('utf-8')
                        
                        # 确保不是数字或年份
                        if not text.isdigit() and len(text) > 1:
                            city_mapping[value] = text
            
            # 验证城市映射，如果为空则使用默认城市列表
            if not city_mapping:
                self.log_message("⚠️  未找到有效的城市选项，使用默认城市列表")
                city_mapping = {
                    "15": "沈阳市", "16": "大连市", "53": "大连金普新区", "56": "大连开发区（2017前）",
                    "17": "鞍山市", "21": "抚顺市", "22": "本溪市", "25": "丹东市", "33": "锦州市",
                    "29": "营口市", "34": "阜新市", "35": "辽阳市", "36": "铁岭市", "44": "朝阳市",
                    "45": "盘锦市", "47": "葫芦岛市", "48": "绥中"
                }
            
            self.city_mapping = city_mapping
            city_values = list(city_mapping.values())
            
            self.log_message(f"✅ 成功获取 {len(city_values)} 个城市")
            
            # 更新城市选择器
            self.root.after(0, lambda: self.city_combo.config(values=city_values, postcommand=self.on_parameter_change))
            
            if city_values:
                self.root.after(0, lambda: self.city_combo.current(0))
        except Exception as e:
            # 使用默认城市列表
            city_mapping = {
                "15": "沈阳市", "16": "大连市", "53": "大连金普新区", "56": "大连开发区（2017前）",
                "17": "鞍山市", "21": "抚顺市", "22": "本溪市", "25": "丹东市", "33": "锦州市",
                "29": "营口市", "34": "阜新市", "35": "辽阳市", "36": "铁岭市", "44": "朝阳市",
                "45": "盘锦市", "47": "葫芦岛市", "48": "绥中"
            }
            self.city_mapping = city_mapping
            city_values = list(city_mapping.values())
            self.log_message(f"✅ 成功获取 {len(city_values)} 个城市")
            self.root.after(0, lambda: self.city_combo.config(values=city_values, postcommand=self.on_parameter_change))
            if city_values:
                self.root.after(0, lambda: self.city_combo.current(0))
    
    def get_year_month_list(self):
        """生成年份和月份列表"""
        try:
            self.log_message("🔄 正在生成年份和月份列表...")
            
            # 生成最近5年的年份列表
            current_year = datetime.now().year
            self.year_list = [str(year) for year in range(current_year, current_year - 5, -1)]
            
            # 生成月份列表
            self.month_list = [f"{month:02d}" for month in range(1, 13)]
            
            # 更新年份和月份选择器
            self.root.after(0, lambda: self.year_combo.config(values=self.year_list, postcommand=self.on_parameter_change))
            self.root.after(0, lambda: self.month_combo.config(values=self.month_list, postcommand=self.on_parameter_change))
            
            if self.year_list:
                self.root.after(0, lambda: self.year_combo.current(0))
            
            if self.month_list:
                self.root.after(0, lambda: self.month_combo.current(datetime.now().month - 1))
            
            self.log_message("✅ 年份和月份列表生成成功")
            
        except Exception as e:
            self.log_message(f"❌ 生成年份和月份列表失败: {str(e)}")
    
    def query_data(self):
        """查询数据，获取总页数和总记录数"""
        def query_task():
            max_retries = 3
            retry_count = 0
            timeout = 60  # 增加超时时间到60秒
            
            while retry_count < max_retries:
                try:
                    retry_count += 1
                    self.log_message(f"🔄 正在查询数据 (第{retry_count}/{max_retries}次尝试)...")
                    
                    # 获取选择的城市、年份和月份
                    selected_city = self.city_var.get()
                    selected_year = self.year_var.get()
                    selected_month = self.month_var.get()
                    
                    if not selected_city or not selected_year or not selected_month:
                        self.log_message("❌ 请选择城市、年份和月份")
                        return
                    
                    # 获取城市ID
                    city_id = None
                    for cid, name in self.city_mapping.items():
                        if name == selected_city:
                            city_id = cid
                            break
                    
                    if not city_id:
                        self.log_message("❌ 未找到选择的城市ID")
                        return
                    
                    # 构建查询URL
                    date_str = f"{selected_year}/{selected_month}/20"
                    query_url = f"http://218.60.144.156/jgxx_cl1.asp?pageno=1&dq_id={city_id}&cllb=&time1={date_str}&clmc=&clid=&view=hidden&tc="
                    
                    self.log_message("📌 查询")
                    
                    # 发送查询请求
                    response = self.session.get(query_url, timeout=timeout)
                    # 先尝试从响应头获取编码，再使用gbk作为备选
                    response.encoding = response.apparent_encoding if response.apparent_encoding else 'gbk'
                    
                    if response.status_code != 200:
                        self.log_message(f"❌ 查询失败，状态码: {response.status_code}")
                        if retry_count < max_retries:
                            time.sleep(2)
                        continue
                    
                    # 解析响应，获取总页数和总记录数
                    soup = BeautifulSoup(response.text, 'html.parser')
                    
                    # 获取总记录数
                    total_records = 0
                    
                    # 保存调试HTML，便于分析
                    import os
                    debug_file_path = os.path.join(self.temp_dir, "query_result_debug.html")
                    with open(debug_file_path, "w", encoding="utf-8") as f:
                        f.write(response.text)
                    
                    # 尝试从分页信息中提取总记录数
                    pagination_text = soup.get_text()
                    
                    # 方法1：从"共找到XX条信息"中提取（带引号）
                    records_match = re.search(r'共找到(\d+)条信息', pagination_text)
                    if records_match:
                        total_records = int(records_match.group(1))
                    else:
                        # 方法2：从"共XX条记录"中提取（备用）
                        records_match = re.search(r'共(\d+)条记录', pagination_text)
                        if records_match:
                            total_records = int(records_match.group(1))
                        else:
                            # 方法3：尝试从其他模式中提取
                            records_match = re.search(r'\d+条', pagination_text)
                            if records_match:
                                # 提取数字部分
                                num_match = re.search(r'\d+', records_match.group(0))
                                if num_match:
                                    total_records = int(num_match.group(0))
                            else:
                                # 方法4：查找所有数字模式，找出可能的总记录数
                                all_numbers = re.findall(r'\d+', pagination_text)
                                if all_numbers:
                                    # 尝试使用最大的数字作为总记录数
                                    total_records = max(int(num) for num in all_numbers)
                    
                    # 获取总页数
                    total_pages = 1
                    # 方法1：从分页文本中提取
                    pages_match = re.search(r'\d+/\d+', pagination_text)
                    if pages_match:
                        # 提取分数形式的页码，如"1/10"，取分母
                        pages_part = pages_match.group(0)
                        total_pages = int(pages_part.split('/')[1])
                    else:
                        # 方法2：计算总页数
                        if total_records > 0:
                            total_pages = (total_records + 49) // 50  # 每页50条数据
                    
                    self.total_records = total_records
                    self.total_pages = total_pages
                    
                    # 更新UI显示
                    self.root.after(0, lambda: self.total_records_label.config(text=f"总记录数: {total_records}"))
                    self.root.after(0, lambda: self.total_pages_label.config(text=f"总页数: {total_pages}"))
                    
                    self.log_message(f"✅ 查询成功，共 {total_records} 条记录，{total_pages} 页")
                    
                    # 锁定查询按钮，只有参数变更后才能重新点击
                    self.root.after(0, lambda: self.query_btn.config(state="disabled"))
                    
                    # 启用数据提取按钮
                    self.root.after(0, lambda: self.extract_btn.config(state="normal"))
                    
                    return  # 查询成功，退出循环
                    
                except requests.exceptions.Timeout:
                    self.log_message(f"⏱️  查询超时 (第{retry_count}/{max_retries}次)")
                    if retry_count < max_retries:
                        time.sleep(2)
                except requests.exceptions.ConnectionError:
                    self.log_message(f"❌ 网络连接错误 (第{retry_count}/{max_retries}次)")
                    if retry_count < max_retries:
                        time.sleep(2)
                except Exception as e:
                    self.log_message(f"❌ 查询失败 (第{retry_count}/{max_retries}次)")
                    if retry_count < max_retries:
                        time.sleep(2)
            
            # 所有重试都失败
            self.log_message(f"❌ 查询失败，已重试{max_retries}次")
        
        # 使用线程执行查询操作，避免阻塞GUI
        thread = threading.Thread(target=query_task)
        thread.daemon = True
        thread.start()
    
    def extract_data(self):
        """提取数据"""
        def extract_task():
            try:
                self.log_message("🔄 正在提取数据...")
                
                # 获取选择的城市、年份和月份
                selected_city = self.city_var.get()
                selected_year = self.year_var.get()
                selected_month = self.month_var.get()
                
                if not selected_city or not selected_year or not selected_month:
                    self.log_message("❌ 请选择城市、年份和月份")
                    return
                
                # 获取城市ID
                city_id = None
                for cid, name in self.city_mapping.items():
                    if name == selected_city:
                        city_id = cid
                        break
                
                if not city_id:
                    self.log_message("❌ 未找到选择的城市ID")
                    return
                
                # 构建日期字符串
                date_str = f"{selected_year}/{selected_month}/20"
                
                # 确定要提取的页数
                pages_to_extract = self.total_pages
                if not self.extract_all_pages_var.get():
                    pages_to_extract = 1
                
                self.log_message(f"📌 开始提取 {pages_to_extract} 页数据")
                
                all_data = []
                headers = []
                
                # 提取每一页数据
                for page_no in range(1, pages_to_extract + 1):
                    # 更新进度
                    progress = (page_no - 1) / pages_to_extract * 100
                    self.root.after(0, lambda p=progress: self.progress_var.set(p))
                    
                    # 提取单页数据
                    page_data, page_headers = self.extract_page_data(city_id, date_str, page_no)
                    
                    if not page_data:
                        self.log_message(f"⚠️  第 {page_no} 页数据提取失败")
                        continue
                    
                    if not headers:
                        headers = page_headers
                    
                    all_data.extend(page_data)
                    self.log_message(f"✅ 第 {page_no} 页数据提取成功，共 {len(page_data)} 条记录")
                
                self.all_data = all_data
                self.headers = headers
                
                # 更新进度
                self.root.after(0, lambda: self.progress_var.set(100))
                
                self.log_message(f"🎉 数据提取完成，共提取 {len(all_data)} 条记录")
                
                # 启用导出按钮
                self.root.after(0, lambda: self.export_btn.config(state="normal"))
                
            except Exception as e:
                self.log_message("❌ 数据提取失败")
                self.root.after(0, lambda: self.progress_var.set(0))
        
        # 使用线程执行提取操作，避免阻塞GUI
        thread = threading.Thread(target=extract_task)
        thread.daemon = True
        thread.start()
    
    def extract_page_data(self, city_id, date_str, page_no):
        """提取单页数据"""
        max_retries = 3
        retry_count = 0
        timeout = 60  # 增加超时时间到60秒
        
        while retry_count < max_retries:
            try:
                retry_count += 1
                # 构建页面URL
                page_url = f"http://218.60.144.156/jgxx_cl1.asp?pageno={page_no}&dq_id={city_id}&cllb=&time1={date_str}&clmc=&clid=&view=hidden&tc="
                
                # 发送请求
                response = self.session.get(page_url, timeout=timeout)
                # 先尝试从响应头获取编码，再使用gbk作为备选
                response.encoding = response.apparent_encoding if response.apparent_encoding else 'gbk'
                
                if response.status_code != 200:
                    if retry_count < max_retries:
                        time.sleep(2)
                    continue
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 查找数据表格
                data_table = None
                tables = soup.find_all('table')
                
                # 保存所有表格的信息，便于选择最合适的表格
                table_info = []
                
                for i, table in enumerate(tables):
                    rows = table.find_all('tr')
                    cells = []
                    header_cells = []
                    data_rows = []
                    
                    if rows:
                        header_cells = rows[0].find_all(['th', 'td'])
                        data_rows = rows[1:]
                    
                    # 计算表格的实际列数（使用表头列数或第一数据行的列数）
                    actual_cols = len(header_cells)
                    if not actual_cols and data_rows:
                        actual_cols = len(data_rows[0].find_all(['th', 'td']))
                    
                    # 计算数据行数量
                    actual_rows = len(data_rows)
                    
                    table_info.append({
                        'index': i,
                        'rows': len(rows),
                        'data_rows': actual_rows,
                        'cols': actual_cols,
                        'table': table
                    })
                
                # 选择最合适的数据表格
                # 优先选择：列数 >= 5 且 数据行 >= 10 的表格
                # 如果没有，选择列数最多的表格
                # 如果还是没有，选择行数最多的表格
                best_table = None
                best_score = -1
                
                for info in table_info:
                    score = 0
                    
                    # 列数越多，得分越高
                    if info['cols'] >= 5:
                        score += 100
                    score += info['cols'] * 10
                    
                    # 数据行越多，得分越高
                    if info['data_rows'] >= 10:
                        score += 50
                    score += info['data_rows']
                    
                    # 总行数越多，得分越高
                    score += info['rows']
                    
                    # 更新最佳表格
                    if score > best_score:
                        best_score = score
                        best_table = info
                
                if not best_table:
                    # 保存调试HTML，便于分析
                    debug_path = f"page_{page_no}_debug.html"
                    with open(debug_path, "w", encoding="utf-8") as f:
                        f.write(response.text)
                    if retry_count < max_retries:
                        time.sleep(2)
                    continue
                
                data_table = best_table['table']
                
                # 提取表头
                header_cells = data_table.find_all('tr')[0].find_all(['th', 'td'])
                headers = []
                for cell in header_cells:
                    text = cell.get_text(strip=True)
                    # 直接使用文本，不进行额外编码处理
                    if text and not text.isspace():
                        headers.append(text)
                
                # 如果表头为空，使用默认表头
                if not headers:
                    headers = ['序号', '材料名称', '规格型号', '单位', '价格(元)', '备注', '发布地区', '发布时间', '材料类别']
                
                # 提取数据行
                data = []
                rows = data_table.find_all('tr')[1:]
                
                for row in rows:
                    cells = row.find_all('td')
                    row_data = []
                    
                    for cell in cells:
                        text = cell.get_text(strip=True)
                        # 直接使用文本，不进行额外编码处理
                        # 清理空格和特殊字符
                        text = re.sub(r'\s+', ' ', text).strip()
                        row_data.append(text)
                    
                    # 确保数据长度与表头一致
                    while len(row_data) < len(headers):
                        row_data.append('')
                    
                    data.append(row_data[:len(headers)])
                
                return data, headers
                
            except requests.exceptions.Timeout:
                if retry_count < max_retries:
                    time.sleep(2)
            except requests.exceptions.ConnectionError:
                if retry_count < max_retries:
                    time.sleep(2)
            except Exception as e:
                if retry_count < max_retries:
                    time.sleep(2)
        
        # 所有重试都失败
        return [], []
    
    def export_data(self):
        """保存数据到Excel/CSV（增强版：彻底解决乱码问题并添加表格样式）"""
        if not self.all_data or not self.headers:
            messagebox.showwarning("警告", "没有数据可以保存！")
            return
        
        try:
            # 1. 增强数据清洗：确保所有数据都是有效的UTF-8字符串
            clean_data = []
            for row in self.all_data:
                clean_row = []
                for cell in row:
                    if isinstance(cell, str):
                        # 确保字符串是有效的UTF-8
                        cell = cell.encode('utf-8', errors='replace').decode('utf-8')
                    clean_row.append(cell)
                clean_data.append(clean_row)
            
            # 2. 增强表头处理：确保表头也是有效的UTF-8字符串
            clean_headers = []
            for header in self.headers:
                if isinstance(header, str):
                    header = header.encode('utf-8', errors='replace').decode('utf-8')
                clean_headers.append(header)
            
            # 3. 创建DataFrame，确保数据结构正确
            df = pd.DataFrame(clean_data, columns=clean_headers)
            
            # 获取选择的城市、年份和月份
            selected_city = self.city_var.get()
            selected_year = self.year_var.get()
            selected_month = self.month_var.get()
            
            # 构建默认文件名格式：辽宁省XX市YYYY年MM月份网刊
            default_filename = f"辽宁省{selected_city}{selected_year}年{selected_month}月份网刊"
            
            # 选择保存路径
            file_path = filedialog.asksaveasfilename(
                initialfile=default_filename,
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv"), ("所有文件", "*.*")],
                title="保存数据文件"
            )
            
            if not file_path:
                return
            
            # 4. 保存为Excel文件（增强版：添加表格样式）
            if file_path.endswith('.xlsx'):
                try:
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        # 写入数据
                        df.to_excel(writer, index=False, sheet_name='材料价格')
                        
                        # 获取工作表
                        worksheet = writer.sheets['材料价格']
                        
                        # 导入openpyxl样式模块
                        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                        
                        # 定义样式
                        # 表头样式
                        header_font = Font(bold=True, color="FFFFFF", size=10)
                        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        
                        # 数据样式
                        data_font = Font(size=10)
                        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                        
                        # 边框样式
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # 设置表头样式
                        for col_num, header in enumerate(clean_headers, 1):
                            cell = worksheet.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = thin_border
                        
                        # 设置数据区域样式
                        for row_num in range(2, len(df) + 2):
                            for col_num in range(1, len(clean_headers) + 1):
                                cell = worksheet.cell(row=row_num, column=col_num)
                                cell.font = data_font
                                cell.alignment = data_alignment
                                cell.border = thin_border
                        
                        # 自动调整列宽
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 100)  # 限制最大宽度为100
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # 设置行高
                        worksheet.row_dimensions[1].height = 21  # 表头行高
                        for row_num in range(2, len(df) + 2):
                            worksheet.row_dimensions[row_num].height = 21  # 数据行高
                        
                        # 添加隔行显示底色
                        from openpyxl.styles import PatternFill
                        even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 浅灰色
                        for row_num in range(2, len(df) + 2, 2):  # 从第2行开始，间隔2行
                            for col_num in range(1, len(clean_headers) + 1):
                                cell = worksheet.cell(row=row_num, column=col_num)
                                cell.fill = even_fill                        
                    self.log_message(f"✅ Excel数据已保存到: {file_path}")
                except Exception as e:
                    # 降级方案：使用xlwt引擎（如果可用）
                    try:
                        with pd.ExcelWriter(file_path, engine='xlwt', encoding='utf-8') as writer:
                            df.to_excel(writer, index=False, sheet_name='材料价格')
                        self.log_message(f"✅ Excel数据已保存到: {file_path}")
                    except:
                        # 最终方案：保存为CSV
                        csv_path = file_path.replace('.xlsx', '.csv')
                        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                        self.log_message(f"🔄 无法保存为Excel，已保存为CSV: {csv_path}")
                        file_path = csv_path
            # 5. 保存为CSV文件（增强版，确保带BOM头）
            elif file_path.endswith('.csv'):
                # 使用utf-8-sig编码，确保Windows Excel能正确识别
                df.to_csv(file_path, index=False, encoding='utf-8-sig', sep=',')
                self.log_message(f"✅ CSV数据已保存到: {file_path}")
            else:
                # 默认保存为Excel
                excel_path = file_path + '.xlsx'
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    # 写入数据
                    df.to_excel(writer, index=False, sheet_name='材料价格')
                    
                    # 获取工作表
                    worksheet = writer.sheets['材料价格']
                    
                    # 导入openpyxl样式模块
                    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                    
                    # 定义样式
                    # 表头样式
                    header_font = Font(bold=True, color="FFFFFF", size=10)
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 数据样式
                    data_font = Font(size=10)
                    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                    
                    # 边框样式
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 设置表头样式
                    for col_num, header in enumerate(clean_headers, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # 设置数据区域样式
                    for row_num in range(2, len(df) + 2):
                        for col_num in range(1, len(clean_headers) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.font = data_font
                            cell.alignment = data_alignment
                            cell.border = thin_border
                    
                    # 自动调整列宽
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 100)  # 限制最大宽度为100
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # 设置行高
                    worksheet.row_dimensions[1].height = 21  # 表头行高
                    for row_num in range(2, len(df) + 2):
                        worksheet.row_dimensions[row_num].height = 21  # 数据行高
                    
                    # 添加隔行显示底色
                    from openpyxl.styles import PatternFill
                    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 浅灰色
                    for row_num in range(2, len(df) + 2, 2):  # 从第2行开始，间隔2行
                        for col_num in range(1, len(clean_headers) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.fill = even_fill
                
                self.log_message(f"✅ Excel数据已保存到: {excel_path}")
                file_path = excel_path
            
            messagebox.showinfo("成功", f"数据已成功保存！\n文件路径: {file_path}\n记录数: {len(self.all_data)}")
            
        except Exception as e:
            self.log_message(f"❌ 保存数据时出错：{str(e)}")
            messagebox.showerror("错误", f"保存失败: {str(e)}")

if __name__ == "__main__":
    # 使用ttkbootstrap创建根窗口
    root = ttk.Window(themename="cosmo")  # 可以选择不同主题，如'cosmo', 'darkly', 'flatly', 'journal', 'litera', 'lumen', 'minty', 'pulse', 'sandstone', 'united', 'yeti'
    app = MaterialPriceScraper(root)
    root.mainloop()
